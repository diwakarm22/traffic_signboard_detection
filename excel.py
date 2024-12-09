from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Initialize the workbook and select the active sheet
wb = Workbook()
ws = wb.active
ws.title = "Trading Journal"

# Define main headers and sub-headers
main_headers = ["Date", "Currency", "Week Bias", "Day Bias", "London Session", "New York Session", "Monthly Report"]
session_sub_headers = ["4hr FVG", "1hr FVG", "Trade", "Profit", "Loss", "RR"]  # Full sub-header list for sessions
daily_report_sub_headers = ["Total Wins", "Total Loss", "Win Percentage", "Total RR"]  # Updated sub-headers for Daily Report

# Font and alignment styles
header_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Define yellow fill color for headers
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Set main headers
ws.merge_cells("A1:A2")  # "Date" spans 2 rows
ws.merge_cells("B1:B2")  # "Currency" spans 2 rows
ws.merge_cells("C1:C2")  # "Week Bias" spans 2 rows
ws.merge_cells("D1:D2")  # "Day Bias" spans 2 rows
ws.merge_cells("E1:J1")  # "London Session" spans 6 columns
ws.merge_cells("K1:P1")  # "New York Session" spans 6 columns
ws.merge_cells("Q1:T1")  # "Daily Report" spans 4 columns

# Write main headers
ws["A1"].value = "Date"
ws["B1"].value = "Currency"
ws["C1"].value = "Week Bias"
ws["D1"].value = "Day Bias"
ws["E1"].value = "London Session"
ws["K1"].value = "New York Session"
ws["Q1"].value = "Daily Report"
for cell in ["A1", "B1", "C1", "D1", "E1", "K1", "Q1"]:
    ws[cell].font = header_font
    ws[cell].alignment = center_alignment
    ws[cell].border = thin_border
    ws[cell].fill = yellow_fill  # Apply yellow fill to header cells

# Write sub-headers for London and New York sessions
for col_num, sub_header in enumerate(session_sub_headers, start=5):
    ws.cell(row=2, column=col_num, value=sub_header).font = header_font
    ws.cell(row=2, column=col_num).alignment = center_alignment
    ws.cell(row=2, column=col_num).border = thin_border
    ws.cell(row=2, column=col_num).fill = yellow_fill  # Apply yellow fill to sub-header cells
    # Repeat for New York Session
    ws.cell(row=2, column=col_num + len(session_sub_headers), value=sub_header).font = header_font
    ws.cell(row=2, column=col_num + len(session_sub_headers)).alignment = center_alignment
    ws.cell(row=2, column=col_num + len(session_sub_headers)).border = thin_border
    ws.cell(row=2, column=col_num + len(session_sub_headers)).fill = yellow_fill  # Apply yellow fill to sub-header cells

# Write sub-headers for Daily Report (only in row 2, columns Q to T)
for col_num, sub_header in enumerate(daily_report_sub_headers, start=17):
    ws.cell(row=2, column=col_num, value=sub_header).font = header_font
    ws.cell(row=2, column=col_num).alignment = center_alignment
    ws.cell(row=2, column=col_num).border = thin_border
    ws.cell(row=2, column=col_num).fill = yellow_fill  # Apply yellow fill to sub-header cells

# Apply borders to the main header cells individually
for col in range(1, 21):  # 1 to 20 (including the new Daily Report columns)
    ws.cell(row=1, column=col).border = thin_border
    ws.cell(row=2, column=col).border = thin_border

# Set column widths for readability
column_widths = [15, 12, 10, 10] + [12] * 12 + [15, 12, 18, 12] # Adjust widths to account for all columns
for col_num, width in enumerate(column_widths, start=1):
    ws.column_dimensions[get_column_letter(col_num)].width = width

# Define currency pairs and the date range
currency_pairs = ["Dxy","EUR/USD", "GBP/USD", "USD/JPY", "USD/CAD", "AUD/USD", "USD/CHF", "NZD/USD",
                  "EUR/GBP", "EUR/JPY", "EUR/CHF", "GBP/JPY", "GBP/CHF", "XAU/USD","XAG/USD","UKO/USD","EUR/AUD","EUR/NZD","EUR/CAD",
                  "GBP/AUD","GBP/NZD" ,"GBP/CAD", "CHF/JPY", "AUD/JPY", "AUD/CHF", "AUD/NZD","AUD/CAD","NZD/JPY","NZD/CHF","NZD/CAD","CAD/JPY","CAD/CHF"]

start_date = datetime.strptime("01-12-2024","%d-%m-%Y")
end_date = datetime.strptime ("12-12-2024","%d-%m-%Y")

# Fill in data for each weekday date and currency pair
row_num = 3  # Start data from the third row
current_date = start_date
while current_date <= end_date:
    # Only fill weekdays (Monday to Friday)
    if current_date.weekday() < 5:  # 0=Monday, 6=Sunday
        first_row_for_date = row_num
        for i, currency in enumerate(currency_pairs):
            # Only write the date in the first row for each group of currencies
            if i == 0:
                ws.cell(row=row_num, column=1, value=current_date.strftime("%d-%m-%Y")).border = thin_border
            else:
                ws.cell(row=row_num, column=1, value="")  # Leave empty with no border

            # Write the currency and apply borders
            ws.cell(row=row_num, column=2, value=currency).border = thin_border

            # Leave Week Bias and Day Bias columns empty for now, but with borders
            ws.cell(row=row_num, column=3, value="").border = thin_border  # Week Bias
            ws.cell(row=row_num, column=4, value="").border = thin_border  # Day Bias

            # Apply borders only to cells where headers would have data
            for col_num in range(5, 21):  # Columns from 5 to 20 (including new Daily Report columns)
                ws.cell(row=row_num, column=col_num, value="").alignment = center_alignment
                # Apply borders only if there's content in header cells
                if col_num < 11 or (col_num >= 11 and col_num < 17):
                    ws.cell(row=row_num, column=col_num).border = thin_border
            row_num += 1
    current_date += timedelta(days=1)

# Add formulas for the Daily Report section (Q3, R3, T3, S3)
ws["Q3"] = "=SUM(H:H)+SUM(N:N)"
ws["R3"] = "=SUM(I:I)+SUM(O:O)"
ws["T3"] = "=SUM(J:J)+SUM(P:P)"
ws["S3"] = "=IF((Q3+R3)>0, Q3 / (Q3 + R3) * 100, 0)"

# Ensure formulas are applied correctly with borders and alignment
for cell in ["Q3", "R3", "T3", "S3"]:
    ws[cell].alignment = center_alignment
    ws[cell].border = thin_border

# Remove borders for cells in Daily Report columns (Q to T) starting from row 4
for row in range(4, row_num):  # Start from row 4 to the last row of filled data
    for col in range(17, 21):  # Columns Q to T
        ws.cell(row=row, column=col).border = None  # Remove border

# Save the workbook after adding the Daily Report columns and formulas
wb.save(r"C:\Users\diwak\Downloads\novemberresult\nov(all).xlsx")